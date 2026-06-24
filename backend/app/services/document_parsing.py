import html
import logging
import os
import re
import zipfile
from io import BytesIO
from pathlib import Path

from defusedxml import ElementTree as ET
from fastapi import HTTPException, status

logger = logging.getLogger(__name__)
SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".html", ".htm", ".docx", ".pdf", ".xlsx", ".xls"}


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def extract_text(filename: str, raw: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"暂不支持 {ext or '未知'} 文件，当前支持：{', '.join(sorted(SUPPORTED_EXTENSIONS))}",
        )
    if ext == ".docx":
        return extract_docx_text(raw)
    if ext == ".pdf":
        return extract_pdf_text(raw)
    if ext in {".xlsx", ".xls"}:
        return extract_xlsx_text(raw)
    text = decode_text(raw)
    if ext in {".html", ".htm"}:
        text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.I)
        text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
        text = re.sub(r"<[^>]+>", " ", text)
        text = html.unescape(text)
    return normalize_text(text)


def extract_pdf_text(raw: bytes) -> str:
    """Extract text from PDF using docling, then pypdf, then raw decode."""
    try:
        from docling.document_converter import DocumentConverter
        from tempfile import NamedTemporaryFile

        with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            converter = DocumentConverter()
            result = converter.convert(tmp_path)
            text = result.document.export_to_markdown()
            return normalize_text(text)
        finally:
            os.unlink(tmp_path)
    except ImportError:
        pass
    except Exception:
        logger.warning("Docling PDF extraction failed, using raw fallback", exc_info=True)

    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(raw))
        pages = []
        for index, page in enumerate(reader.pages, 1):
            page_text = page.extract_text() or ""
            if page_text.strip():
                pages.append(f"--- Page {index} ---\n{page_text}")
        if pages:
            return normalize_text("\n\n".join(pages))
    except Exception:
        logger.warning("pypdf extraction failed, using raw PDF fallback", exc_info=True)

    text = decode_text(raw)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', text)
    return normalize_text(text)


def extract_xlsx_text(raw: bytes) -> str:
    """Extract text from Excel using docling (with fallback to openpyxl)."""
    try:
        from docling.document_converter import DocumentConverter
        from tempfile import NamedTemporaryFile

        suffix = ".xlsx"
        with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            converter = DocumentConverter()
            result = converter.convert(tmp_path)
            text = result.document.export_to_markdown()
            return normalize_text(text)
        finally:
            os.unlink(tmp_path)
    except ImportError:
        pass
    except Exception:
        logger.warning("Docling spreadsheet extraction failed, using openpyxl fallback", exc_info=True)

    # Fallback: openpyxl
    try:
        from openpyxl import load_workbook

        value_wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        formula_wb = load_workbook(BytesIO(raw), read_only=True, data_only=False)
        lines = []
        for sheet_name in value_wb.sheetnames:
            ws = value_wb[sheet_name]
            formula_ws = formula_wb[sheet_name]
            lines.append(f"--- Sheet: {sheet_name} ---")
            lines.append(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            for row_index, row in enumerate(ws.iter_rows(values_only=True), 1):
                formula_row = next(formula_ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
                cells = []
                for value, formula in zip(row, formula_row):
                    if value is None and formula is None:
                        cells.append("")
                    elif isinstance(formula, str) and formula.startswith("="):
                        cells.append(f"{value if value is not None else ''} ({formula})")
                    else:
                        cells.append(str(value if value is not None else formula))
                line = "\t".join(cells)
                if line.strip():
                    lines.append(line)
        value_wb.close()
        formula_wb.close()
        return normalize_text("\n".join(lines))
    except Exception:
        logger.warning("openpyxl extraction failed, using raw spreadsheet fallback", exc_info=True)
        return normalize_text(decode_text(raw))


def decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def extract_docx_text(raw: bytes) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(BytesIO(raw)) as archive:
        document = archive.read("word/document.xml")
    root = ET.fromstring(document)
    lines = []
    for paragraph in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            lines.append(line)
    return normalize_text("\n".join(lines))


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

